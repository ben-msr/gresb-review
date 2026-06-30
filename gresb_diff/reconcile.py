"""Reconcile flagged like-for-like differences against the GRESB asset-level
Excel: for each mismatch, identify which asset(s) explain it and classify why.

A flagged matrix line is a portfolio/property-type total. The Absolute total is
the sum over all assets (ownership-weighted); the GRESB and Word like-for-like
figures are subsets. The gap between them is explained by one of:
  * word_dropped   - asset(s) GRESB keeps in like-for-like but the Word doc drops
  * word_added     - asset(s)/value the Word doc includes that GRESB does not
  * negatives      - negative asset values (GRESB floors them to 0; Word keeps raw)
  * gresb_adjustment - no asset subset explains it and the Word figure matches the
                       source data, i.e. a GRESB-side estimate/gross-up
Only EN1/GH1/WT1 consumption + Floor-Area-Covered lines are reconciled here;
EN1FA/EN1RE/WS1/BC are out of scope (separate tables / known issues).
"""
from __future__ import annotations

import io
import itertools

# canonical zone -> energy/water area code; fuel -> suffix; GHG scope -> code
_AREA = {
    "WB_LANDLORD": "w", "WB_TENANT": "w", "COMMON": "lc_bc", "SHARED": "lc_bs",
    "TS_LANDLORD": "lc_t", "TS_TENANT": "tc_t", "EXT_LANDLORD": "lc_o",
    "EXT_TENANT": "tc_o",
}
_FUEL = {"ELECTRIC": "e", "FUEL": "f", "DISTRICT": "d"}
_SCOPE = {"SCOPE_1": "s1", "SCOPE_2": "s2_lb", "SCOPE_3": "s3"}
# metric token -> (data_year, abs|cov). Energy consumption is kWh -> MWh (x0.001).
_METRIC = {"prior": (2024, "abs"), "reporting": (2025, "abs"),
           "fa_covered": (2025, "cov")}


def column_for(question, zone, fuel_or_scope, kind):
    """(sheet, column_code, scale) for a canonical row, or None if unmapped."""
    if question == "EN1":
        area = _AREA.get(zone)
        fuel = _FUEL.get(fuel_or_scope)
        if area is None or fuel is None:
            return None
        return "Energy", f"en_{kind}_{area}{fuel}", (0.001 if kind == "abs" else 1.0)
    if question == "WT1":
        if zone.startswith("WB"):
            return "Water", f"wat_{kind}_w", 1.0
        area = _AREA.get(zone)
        return ("Water", f"wat_{kind}_{area}", 1.0) if area else None
    if question == "GH1":
        scope = _SCOPE.get(fuel_or_scope)
        if scope is None:
            return None
        z = "w" if zone == "WB" else "o"
        return "GHG", f"ghg_{kind}_{scope}_{z}", 1.0
    return None


class AssetData:
    """Per-asset GRESB asset-level workbook: ownership and metric columns."""

    def __init__(self, xlsx_bytes):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True,
                                    data_only=True)
        self.ownership = {}
        for r in wb["Asset Characteristics"].iter_rows(min_row=2, values_only=True):
            name, ptc, o = r[2], r[4], r[10]
            if name and isinstance(o, (int, float)) and ptc and ":" in str(ptc):
                self.ownership[name] = o / 100.0
        self.sheets = {}
        for s in ("Energy", "GHG", "Water"):
            if s not in wb.sheetnames:
                continue
            ws = wb[s]
            hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            hm = {str(v): i for i, v in enumerate(hdr) if v}
            rows = [(r[1], str(r[2]), int(r[3]), r)
                    for r in ws.iter_rows(min_row=2, values_only=True)
                    if r[1] and r[2] and isinstance(r[3], (int, float))]
            self.sheets[s] = (hm, rows)
        wb.close()

    def column(self, sheet, code, year, ptcode, scale):
        """{asset_name: ownership-weighted value} for one column/year/property
        type, or None if the column or sheet is absent."""
        if sheet not in self.sheets:
            return None
        hm, rows = self.sheets[sheet]
        i = hm.get(code)
        if i is None:
            return None
        out = {}
        for name, ptc, yr, r in rows:
            if ptc == ptcode and yr == year:
                v = r[i]
                if isinstance(v, (int, float)) and v:
                    out[name] = v * scale * self.ownership.get(name, 1.0)
        return out


def _subsets(pool, target, max_k=4):
    """Asset sets (size <= max_k) whose values sum to `target`. [] = target is
    ~0; None = no exact subset found."""
    tol = max(1.0, abs(target) * 0.002)
    if abs(target) < tol:
        return []
    names = list(pool)
    for k in range(1, min(max_k, len(names)) + 1):
        hits = [set(c) for c in itertools.combinations(names, k)
                if abs(sum(pool[x] for x in c) - target) < tol]
        if hits:
            return hits
    return None


_ZONE_LABEL = {
    "WB_LANDLORD": "Whole Building (Landlord)", "WB_TENANT": "Whole Building (Tenant)",
    "WB": "Whole Building", "EXT": "Exterior",
    "COMMON": "Common Areas", "SHARED": "Shared Services",
    "TS_LANDLORD": "Tenant Space (Landlord)", "TS_TENANT": "Tenant Space (Tenant)",
    "EXT_LANDLORD": "Exterior (Landlord)", "EXT_TENANT": "Exterior (Tenant)",
}
_FS_LABEL = {"ELECTRIC": "Electric", "FUEL": "Fuel", "DISTRICT": "District",
             "WATER": "Water", "SCOPE_1": "Scope 1", "SCOPE_2": "Scope 2",
             "SCOPE_3": "Scope 3"}


def _row_label(zone, fs):
    z = _ZONE_LABEL.get(zone, zone)
    return f"{z} {_FS_LABEL[fs]}" if fs in _FS_LABEL else z


def reconcile_one(asset, diff):
    """Reconcile a single flagged Difference. Returns a dict (category + driver
    assets + values) or None if the row is out of scope / unmappable."""
    parts = diff.canonical_id.split(".")
    if parts[0] not in ("EN1", "GH1", "WT1") or len(parts) < 4:
        return None
    question, zonefuel, metric = parts[0], parts[1], parts[3]
    if metric not in _METRIC:
        return None
    zone, _, fs = zonefuel.partition("|")
    year, kind = _METRIC[metric]
    mapped = column_for(question, zone, fs, kind)
    if mapped is None:
        return None
    sheet, code, scale = mapped
    ptcode = diff.property_type.rsplit(" | ", 1)[0]
    pool = asset.column(sheet, code, year, ptcode, scale)
    base = {"canonical_id": diff.canonical_id, "property_type": diff.property_type,
            "metric": metric, "question": question, "section": diff.section,
            "row_label": _row_label(zone, fs)}
    if pool is None:
        return {**base, "category": "no_data", "detail": f"column {code} not found"}
    if not pool:
        return {**base, "category": "no_data",
                "detail": f"no asset values for {ptcode} {year}"}
    try:
        pdf = float(str(diff.pdf_value).replace(",", ""))
        docx = float(str(diff.docx_value).replace(",", ""))
    except (TypeError, ValueError):
        return None
    gap = round(pdf - docx, 2)
    pool_total = round(sum(pool.values()), 2)
    ttol = max(1.0, abs(pool_total) * 0.003)
    result = {**base, "pdf_value": pdf, "docx_value": docx, "gap": gap,
              "pool_total": pool_total}
    # Negatives are checked first: they break the aggregation (GRESB floors them
    # to 0; the Word doc carries them raw) and also make pool_total unreliable.
    negatives = sorted(n for n, v in pool.items() if v < 0)
    if negatives:
        return {**result, "category": "negatives", "assets": negatives}
    # A like-for-like figure can't exceed the asset-data total; if it does, the
    # difference isn't an included/excluded asset.
    if pdf > pool_total + ttol:
        # GRESB above the source total -> GRESB-side estimate/gross-up; the Word
        # figure matches the data.
        return {**result, "category": "gresb_adjustment", "assets": []}
    if docx > pool_total + ttol:
        # Word above the source total -> Word includes something not in the data.
        return {**result, "category": "word_extra", "assets": []}
    sub = _subsets(pool, abs(gap))
    if sub:
        return {**result, "assets": sorted(next(iter(sub))),
                "ambiguous": len(sub) > 1,
                "category": "word_dropped" if gap > 0 else "word_added"}
    return {**result, "category": "gresb_adjustment", "assets": []}


def reconcile(asset, compare_result):
    """Reconcile every in-scope flagged difference; returns a list of dicts."""
    out = []
    for d in compare_result.differences:
        r = reconcile_one(asset, d)
        if r is not None:
            out.append(r)
    return out


_CATEGORY_LABEL = {
    "word_dropped": "Word doc dropped from like-for-like",
    "word_added": "Word doc included (GRESB excluded from like-for-like)",
    "word_extra": "Word doc value exceeds the asset data — includes extra",
    "negatives": "Negative asset value — GRESB floors to 0, Word keeps it raw",
    "gresb_adjustment": "GRESB-side adjustment — Word matches the asset data",
    "no_data": "No matching asset-level data",
}
# Priority when a row's metrics resolve to different categories (most specific
# / actionable first).
_CATEGORY_RANK = ["word_dropped", "word_added", "negatives", "word_extra",
                  "gresb_adjustment", "no_data"]


def reconcile_report(recons) -> str:
    """Markdown for the reconciliation section: grouped by '<CODE> <Section> -
    <Property Type>', one line per flagged row giving the cause and driver
    asset(s). Empty string if there is nothing to show."""
    from collections import OrderedDict
    groups: "OrderedDict" = OrderedDict()
    for r in recons:
        head = f"{r['question']} {r['section']} - {r['property_type']}"
        row = groups.setdefault(head, OrderedDict()).setdefault(
            r["row_label"], {"cats": set(), "assets": [], "ambiguous": False})
        row["cats"].add(r["category"])
        row["ambiguous"] = row["ambiguous"] or r.get("ambiguous", False)
        for a in r.get("assets", []):
            if a not in row["assets"]:
                row["assets"].append(a)

    lines = []
    for head, rows in groups.items():
        lines.append(f"**{head}**")
        for row_label, info in rows.items():
            cat = next((c for c in _CATEGORY_RANK if c in info["cats"]), "no_data")
            label = _CATEGORY_LABEL.get(cat, cat)
            line = f"- {row_label} — {label}"
            if info["assets"]:
                line += ": " + ", ".join(info["assets"])
                if info["ambiguous"]:
                    line += " _(one of several possible sets)_"
            lines.append(line)
        lines.append("")
    return "\n".join(lines).strip()
